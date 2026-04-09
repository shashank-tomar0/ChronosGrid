import openpyxl
import json

wb = openpyxl.load_workbook(r'Individual TT Format.xlsx')
print(f"Sheet names: {wb.sheetnames}")
print(f"Total sheets: {len(wb.sheetnames)}")
print("="*80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n--- Sheet: '{sheet_name}' (Rows: {ws.max_row}, Cols: {ws.max_column}) ---")
    # Print all data
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=False):
        row_data = []
        for cell in row:
            val = cell.value
            if val is not None:
                row_data.append(f"[{cell.coordinate}]={val}")
        if row_data:
            print("  ".join(row_data))
    if ws.max_row > 20:
        print(f"  ... ({ws.max_row - 20} more rows)")
    print()

# Also try to get merged cells info
for sheet_name in wb.sheetnames[:5]:
    ws = wb[sheet_name]
    if ws.merged_cells.ranges:
        print(f"Merged cells in '{sheet_name}': {list(ws.merged_cells.ranges)}")

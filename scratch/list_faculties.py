import openpyxl

wb = openpyxl.load_workbook('TM-08 (HACKWITHINFY).xlsx', data_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
for name in wb.sheetnames:
    ws = wb[name]
    faculty = str(ws['A7'].value or '').replace('Name of the Faculty:', '').strip()
    print(f"Sheet: {name} -> Faculty: {faculty}")

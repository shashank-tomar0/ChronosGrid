import openpyxl

wb = openpyxl.load_workbook('TM-08 (HACKWITHINFY).xlsx', data_only=True)
sheets = ['Meenakshi', 'Ayush', 'Vishu', 'Manish', 'Yash']

for s in sheets:
    ws = wb[s]
    a7 = ws['A7'].value
    a8 = ws['A8'].value
    print(f"Sheet: {s}")
    print(f"  A7: {a7}")
    print(f"  A8: {a8}")
    # Also check nearby cells
    for r in range(7, 9):
        for c in range(1, 6):
            val = ws.cell(row=r, column=c).value
            if val:
                print(f"  Cell {ws.cell(row=r, column=c).coordinate}: {val}")
    print("-" * 20)

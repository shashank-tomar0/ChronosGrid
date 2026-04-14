import openpyxl

wb = openpyxl.load_workbook('TM-08 (HACKWITHINFY).xlsx', data_only=True)
days = {10: 'MONDAY', 14: 'TUESDAY', 18: 'WEDNESDAY', 22: 'THURSDAY', 26: 'FRIDAY'}
col_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

potential_ambiguities = []

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for start_row, day_name in days.items():
        for col in col_letters:
            # Check all 4 rows for this day and column
            values = []
            for r in range(start_row, start_row + 4):
                val = ws[f'{col}{r}'].value
                if val:
                    values.append((r, val))
            
            if len(values) > 1:
                potential_ambiguities.append({
                    'sheet': sheet_name,
                    'day': day_name,
                    'col': col,
                    'values': values
                })

if potential_ambiguities:
    print(f"Found {len(potential_ambiguities)} instances where multiple rows in a day/slot block have values.")
    for amb in potential_ambiguities[:10]:
        print(f"Sheet: {amb['sheet']}, Day: {amb['day']}, Col: {amb['col']}")
        for r, v in amb['values']:
            print(f"  Row {r}: {v}")
else:
    print("No instances found where multiple rows in a day/slot block have values (within rows 0-3 of each day).")

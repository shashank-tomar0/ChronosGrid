import openpyxl

wb = openpyxl.load_workbook('TM-08 (HACKWITHINFY).xlsx', data_only=True)
sheets_to_check = ['Gargi', 'Meenakshi', 'Shyam', 'Yash', 'Anshuli', 'Sheet1']

print("DEBUGGING EXCEL EXTRACTION")
print("="*40)

for sname in sheets_to_check:
    if sname in wb.sheetnames:
        ws = wb[sname]
        a7 = ws['A7'].value
        a8 = ws['A8'].value
        print(f"Sheet: {sname}")
        print(f"  A7: {repr(a7)}")
        print(f"  A8: {repr(a8)}")
        print("-" * 20)
    else:
        print(f"Sheet: {sname} NOT FOUND")

# Also list some sheets at the beginning to see if they differ
print("\nFIRST 5 SHEETS:")
for i in range(min(5, len(wb.sheetnames))):
    sname = wb.sheetnames[i]
    ws = wb[sname]
    print(f"Sheet: {sname} | A7: {ws['A7'].value}")

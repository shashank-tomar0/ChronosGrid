import openpyxl
import json

wb = openpyxl.load_workbook(r'TM-08 (HACKWITHINFY).xlsx', data_only=True)
print(f"Total sheets: {len(wb.sheetnames)}")
print(f"Sheet names: {wb.sheetnames}")
print("="*80)

# The structure is consistent across sheets:
# Row 7: Name of Faculty
# Row 8: Department, L/T/P counts
# Row 9: Time slots header
# Row 10-13: MONDAY (merged A10:A13)
# Row 14-17: TUESDAY (merged A14:A17)
# Row 18-21: WEDNESDAY (merged A18:A21)
# Row 22-25: THURSDAY (merged A22:A25)
# Row 26-29: FRIDAY (merged A26:A29)
# Columns B-J = 9 time periods

# Let's extract all teacher timetable data
days = {10: 'MONDAY', 14: 'TUESDAY', 18: 'WEDNESDAY', 22: 'THURSDAY', 26: 'FRIDAY'}
time_slots = ['08:50-09:40', '09:40-10:30', '10:40-11:30', '11:30-12:20', 
              '12:20-01:10', '01:10-02:00', '02:00-02:50', '02:50-03:40', '03:40-04:30']
# Columns B through J
col_letters = ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']

all_teachers = {}

for sheet_name in wb.sheetnames:
    if sheet_name == 'Sheet1':
        continue
        
    ws = wb[sheet_name]
    
    # Get faculty name from A7
    faculty_raw = str(ws['A7'].value or '').replace('Name of the Faculty:', '').strip()
    
    # If faculty_raw is empty and sheet_name is Sheet1-like, skip it
    if not faculty_raw and (sheet_name.startswith('Sheet') or len(sheet_name) < 2):
        continue
    
    # Handle cases like "Ms." or "Mr." or empty by using sheet name
    if faculty_raw.lower() in ['ms.', 'mr.', 'dr.', ''] or len(faculty_raw) <= 3:
        faculty_name = f"{faculty_raw} {sheet_name}".strip()
    else:
        faculty_name = faculty_raw
    
    # Clean up name (e.g., remove common parenthetical student info if it clutters the name)
    # But for now, just ensure it's not JUST a title.
    
    dept = str(ws['A8'].value or '').replace('Department:', '').strip()
    
    # Get L/T/P counts
    l_count = ws['G8'].value
    t_count = ws['H8'].value  
    p_count = ws['I8'].value
    
    print(f"\nTeacher: {faculty_name} | Dept: {dept} | L={l_count}, T={t_count}, P={p_count}")
    
    teacher_schedule = {}
    for day_row, day_name in days.items():
        day_schedule = {}
        for idx, col in enumerate(col_letters):
            # Concatenate non-empty values from all 4 rows in the day's block
            cell_values = []
            for r in range(day_row, day_row + 4):
                val = ws[f'{col}{r}'].value
                if val and str(val).strip() and str(val).strip() != '':
                    cell_values.append(str(val).strip())
            
            if cell_values:
                # Join with newline to maintain structure, or space if preferred
                day_schedule[time_slots[idx]] = "\n".join(cell_values)
        
        teacher_schedule[day_name] = day_schedule
        
        # Print schedule
        busy_slots = [f"  {slot}: {subj}" for slot, subj in day_schedule.items()]
        free_slots = [time_slots[i] for i in range(9) if time_slots[i] not in day_schedule]
        if busy_slots:
            print(f"  {day_name}: {len(day_schedule)} classes, {len(free_slots)} free")
            for b in busy_slots:
                print(f"    {b}")
        else:
            print(f"  {day_name}: OFF (all 9 slots free)")
    
    all_teachers[sheet_name] = {
        'name': faculty_name,
        'department': dept,
        'lectures': l_count,
        'tutorials': t_count,
        'practicals': p_count,
        'schedule': teacher_schedule
    }

# Save as JSON for later use
# Save as JSON for later use
with open('timetable_data.json', 'w') as f:
    json.dump(all_teachers, f, indent=2)

# Also save to the Next.js project directory
try:
    with open('time-table-next/src/lib/timetable_data.json', 'w') as f:
        json.dump(all_teachers, f, indent=2)
    print("Also synced to time-table-next/src/lib/timetable_data.json")
except Exception as e:
    print(f"Warning: Could not sync to Next.js directory: {e}")

print(f"\n\nSaved data for {len(all_teachers)} teachers.")

